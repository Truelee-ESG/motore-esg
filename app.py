import os
import re
from datetime import datetime
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
import pytesseract

def seleziona_cartella():
    path = filedialog.askdirectory()
    if path:
        entry_path.delete(0, tk.END)
        entry_path.insert(0, path)

def analizza_file_comuni(cartella):
    if not os.path.exists(cartella):
        messagebox.showerror("Errore", "La cartella specificata non esiste.")
        return None
    valid_extensions = ('.pdf', '.jpg', '.jpeg', '.png')
    file_list = []
    for root_dir, _, files in os.walk(cartella):
        for filename in files:
            if filename.lower().endswith(valid_extensions):
                file_list.append((root_dir, filename))
    return file_list

def analizza_file_trasporti(cartella):
    if not os.path.exists(cartella):
        messagebox.showerror("Errore", "La cartella specificata non esiste.")
        return None
    
    valid_extensions = ('.pdf', '.jpg', '.jpeg', '.png')
    target_folder_keywords = ['gasolio', 'benzina', 'mezzi', 'auto', 'automezzi', 'trasporti', 'carburante', 'fleet', 'veicoli']
    
    file_list = []
    for root_dir, _, files in os.walk(cartella):
        folder_name = os.path.basename(root_dir).lower()
        is_target_dir = any(kw in folder_name or kw in root_dir.lower() for kw in target_folder_keywords)
        
        if is_target_dir:
            for filename in files:
                if filename.lower().endswith(valid_extensions):
                    file_list.append((root_dir, filename))
                    
    if not file_list:
        for root_dir, _, files in os.walk(cartella):
            for filename in files:
                if filename.lower().endswith(valid_extensions):
                    file_list.append((root_dir, filename))
                    
    return file_list

def estrai_mesi_e_anno(text, text_lower, filename=""):
    mesi_mappa = {
        'gennaio': 'Gennaio', 'febbraio': 'Febbraio', 'marzo': 'Marzo', 'aprile': 'Aprile',
        'maggio': 'Maggio', 'giugno': 'Giugno', 'luglio': 'Luglio', 'agosto': 'Agosto',
        'settembre': 'Settembre', 'ottobre': 'Ottobre', 'novembre': 'Novembre', 'dicembre': 'Dicembre',
        'gen': 'Gen', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr',
        'mag': 'Mag', 'giu': 'Giu', 'lug': 'Lug', 'ago': 'Ago',
        'set': 'Set', 'ott': 'Ott', 'nov': 'Nov', 'dic': 'Dic'
    }
    anno = "Non rilevato"
    periodo = "Non rilevato"
    
    source_to_check = f"{filename} {text}"
    years = re.findall(r'\b(20\d{2})\b', source_to_check)
    if years:
        anno = years[0]
        
    found_mesi = []
    source_lower = f"{filename.lower()} {text_lower}"
    for m_key, m_val in mesi_mappa.items():
        if re.search(r'\b' + m_key + r'\b', source_lower):
            if m_val not in found_mesi:
                found_mesi.append(m_val)
    if found_mesi:
        periodo = found_mesi[0] if len(found_mesi) == 1 else f"{found_mesi[0]} - {found_mesi[-1]}"
        
    return periodo, anno

def estrai_dati_trasporto_avanzato(page, text, text_lower):
    quantita = "Non rilevato"
    unita_misura = "Non rilevato"
    
    try:
        words = page.extract_words(extra_attrs=["size"])
        # Cerca etichette di quantità
        for i, word in enumerate(words):
            w_text = word['text'].lower()
            if any(k in w_text for k in ['quantità', 'quantita', 'q.tà', 'q,tà', 'qta', 'um', 'u.m.']):
                # Analizza le parole vicine (a destra o sotto nella lettura geometrica)
                for j in range(i + 1, min(len(words), i + 8)):
                    candidate_word = words[j]['text']
                    clean_c = candidate_word.replace('.', '').replace(',', '.')
                    if re.match(r'^\d+[\.,]?\d*$', clean_c) and quantita == "Non rilevato":
                        quantita = candidate_word
                    elif any(u in candidate_word.lower() for u in ['l', 'litri', 'litro', 'kg', 'kg.', 'kilogrammi']):
                        u_low = candidate_word.lower()
                        if 'kg' in u_low or 'kilogrammi' in u_low:
                            unita_misura = "kg"
                        else:
                            unita_misura = "Litri"
    except Exception:
        pass

    # Fallback tramite Regex strutturata sul testo se l'analisi geometrica non basta
    if quantita == "Non rilevato":
        match_q = re.search(r'(?:quantit[aà]|q[\.,]tà|qta)\D{0,20}(\d{1,3}(?:\.\d{3})*[\.,]?\d*)', text_lower)
        if match_q:
            quantita = match_q.group(1)

    if unita_misura == "Non rilevato":
        if re.search(r'\b(?:litri|litro|\bL\b)\b', text_lower):
            unita_misura = "Litri"
        elif re.search(r'\b(?:kg|kilogrammi)\b', text_lower):
            unita_misura = "kg"

    return quantita, unita_misura

def avvia_energia_elettrica():
    threading.Thread(target=_process_energia_elettrica, daemon=True).start()

def _process_energia_elettrica():
    azienda = entry_azienda.get().strip()
    cartella = entry_path.get().strip()
    if not azienda or not cartella:
        messagebox.showerror("Errore", "Inserisci il nome dell'azienda e seleziona una cartella valida.")
        return

    file_list = analizza_file_comuni(cartella)
    if not file_list:
        messagebox.showwarning("Attenzione", "Nessun file valido trovato nella cartella.")
        return

    total_files = len(file_list)
    progress_elec.config(maximum=total_files, value=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumi Elettrici"
    
    ws.append([f"Estrazione dati bollette energia elettrica - {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.cell(row=1, column=1).font = Font(size=12, bold=True, color="1F497D")
    ws.append([])
    
    headers = ["Azienda", "Nome File", "Periodo", "Anno", "Quantità", "Unità di misura"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    valid_count = 0
    for idx, (root_dir, filename) in enumerate(file_list, 1):
        file_path = os.path.join(root_dir, filename)
        ext = filename.lower()
        quantita = "Non rilevato"
        unita_misura = "kWh"
        periodo = "Non rilevato"
        anno = "Non rilevato"
        text = ""
        
        try:
            if ext.endswith('.pdf'):
                with pdfplumber.open(file_path) as pdf:
                    if len(pdf.pages) > 0:
                        page = pdf.pages[0]
                        text = page.extract_text() or ""
                        text_lower = text.lower()
                        periodo, anno = estrai_mesi_e_anno(text, text_lower, filename)
                        
                        if "energia elettrica" in text_lower or "kwh" in text_lower:
                            try:
                                words = page.extract_words(extra_attrs=["size"])
                                candidates = []
                                for i, word in enumerate(words):
                                    w_text = word['text']
                                    if re.match(r'^(?:kWh|KWh|KWH)$', w_text):
                                        for j in range(max(0, i-3), i):
                                            prev_word = words[j]['text']
                                            clean_prev = prev_word.replace('.', '').replace(',', '.')
                                            if re.match(r'^\d+[\.,]?\d*$', clean_prev):
                                                is_band = False
                                                for k in range(max(0, j-2), min(len(words), j+3)):
                                                    if re.match(r'^F[1-3]$', words[k]['text'], re.IGNORECASE):
                                                        is_band = True
                                                        break
                                                if not is_band:
                                                    candidates.append({
                                                        'valore': prev_word,
                                                        'size': words[j].get('size', 0)
                                                    })
                                if candidates:
                                    candidates.sort(key=lambda x: x['size'], reverse=True)
                                    quantita = candidates[0]['valore']
                                else:
                                    match_kwh = re.search(r'(\d+[\.,]?\d*)\s*(?:kWh|KWh|KWH)', text)
                                    if match_kwh:
                                        quantita = match_kwh.group(1)
                            except Exception:
                                match_kwh = re.search(r'(\d+[\.,]?\d*)\s*(?:kWh|KWh|KWH)', text)
                                if match_kwh:
                                    quantita = match_kwh.group(1)
            elif ext.endswith(('.jpg', '.jpeg', '.png')):
                text = pytesseract.image_to_string(Image.open(file_path)) or ""
                text_lower = text.lower()
                periodo, anno = estrai_mesi_e_anno(text, text_lower, filename)
                match_kwh = re.search(r'(\d+[\.,]?\d*)\s*(?:kWh|KWh|KWH)', text)
                if match_kwh:
                    quantita = match_kwh.group(1)

            if quantita != "Non rilevato":
                row_idx = ws.max_row + 1
                ws.append([azienda, filename, periodo, anno, quantita, unita_misura])
                ws.cell(row=row_idx, column=2).hyperlink = os.path.abspath(file_path)
                ws.cell(row=row_idx, column=2).font = Font(color="0563C1", underline="single")
                valid_count += 1
        except Exception:
            pass

        progress_elec.config(value=idx)
        lbl_status_elec.config(text=f"Controllati: {idx} / {total_files} (Validi: {valid_count})")

    if valid_count == 0:
        messagebox.showwarning("Attenzione", "Nessuna bolletta elettrica valida trovata.")
        return

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    out_path = os.path.join(desktop, f"Consumi_Elettrici_{azienda.replace(' ', '_')}.xlsx")
    wb.save(out_path)
    messagebox.showinfo("Successo", f"File Excel salvato sul Desktop:\n{os.path.basename(out_path)}")

def avvia_trasporti():
    threading.Thread(target=_process_trasporti, daemon=True).start()

def _process_trasporti():
    azienda = entry_azienda.get().strip()
    cartella = entry_path.get().strip()
    if not azienda or not cartella:
        messagebox.showerror("Errore", "Inserisci il nome dell'azienda e seleziona una cartella valida.")
        return

    file_list = analizza_file_trasporti(cartella)
    if not file_list:
        messagebox.showwarning("Attenzione", "Nessun file valido trovato nelle cartelle di trasporto.")
        return

    total_files = len(file_list)
    progress_trans.config(maximum=total_files, value=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trasporti Carburante"
    
    ws.append([f"Estrazione dati fatture carburante - {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.cell(row=1, column=1).font = Font(size=12, bold=True, color="2E7D32")
    ws.append([])
    
    headers = ["Azienda", "Nome File", "Periodo", "Anno", "Quantità", "Unità di misura", "Carburante"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    valid_count = 0
    for idx, (root_dir, filename) in enumerate(file_list, 1):
        file_path = os.path.join(root_dir, filename)
        abs_path = os.path.abspath(file_path)
        ext = filename.lower()
        
        quantita = "Non rilevato"
        unita_misura = "Non rilevato"
        carburante = "Non rilevato"
        periodo = "Non rilevato"
        anno = "Non rilevato"
        text = ""

        try:
            if ext.endswith('.pdf'):
                with pdfplumber.open(file_path) as pdf:
                    if len(pdf.pages) > 0:
                        page = pdf.pages[0]
                        text = page.extract_text() or ""
                        text_lower = text.lower()
                        
                        periodo, anno = estrai_mesi_e_anno(text, text_lower, filename)

                        combined_check = f"{root_dir.lower()} {text_lower}"
                        if "benzina" in combined_check:
                            carburante = "Benzina"
                        elif any(k in combined_check for k in ["gasolio", "diesel", "carbur"]):
                            carburante = "Diesel"

                        quantita, unita_misura = estrai_dati_trasporto_avanzato(page, text, text_lower)

            elif ext.endswith(('.jpg', '.jpeg', '.png')):
                text = pytesseract.image_to_string(Image.open(file_path)) or ""
                text_lower = text.lower()
                periodo, anno = estrai_mesi_e_anno(text, text_lower, filename)
                
                combined_check = f"{root_dir.lower()} {text_lower}"
                if "benzina" in combined_check:
                    carburante = "Benzina"
                elif any(k in combined_check for k in ["gasolio", "diesel", "carbur"]):
                    carburante = "Diesel"

                match_qty = re.search(r'(\d{1,3}(?:\.\d{3})*[\.,]?\d*)\s*(?:litri|Litri|L\b|litro|kg|KG)', text, re.IGNORECASE)
                if match_qty:
                    quantita = match_qty.group(1)
                    unita_misura = "kg" if "kg" in match_qty.group(0).lower() else "Litri"

            row_idx = ws.max_row + 1
            ws.append([azienda, filename, periodo, anno, quantita, unita_misura, carburante])
            
            cell_file = ws.cell(row=row_idx, column=2)
            cell_file.hyperlink = abs_path
            cell_file.font = Font(color="0563C1", underline="single")
            
            valid_count += 1
        except Exception:
            row_idx = ws.max_row + 1
            ws.append([azienda, filename, "-", "-", "Non rilevato", "Non rilevato", "Non rilevato"])
            cell_file = ws.cell(row=row_idx, column=2)
            cell_file.hyperlink = abs_path
            cell_file.font = Font(color="0563C1", underline="single")
            valid_count += 1

        progress_trans.config(value=idx)
        lbl_status_trans.config(text=f"Processati: {idx} / {total_files}")

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    out_path = os.path.join(desktop, f"Consumi_Trasporti_{azienda.replace(' ', '_')}.xlsx")
    wb.save(out_path)
    messagebox.showinfo("Successo", f"File Excel salvato sul Desktop:\n{os.path.basename(out_path)}")

# Configurazione Interfaccia Grafica (Layout Motore ESG)
root = tk.Tk()
root.title("Motore ESG - Estrazione Consumi")
root.geometry("740x600")
root.resizable(False, False)
root.configure(bg="#ffffff")

FONT_FAMILY = "Segoe UI"

lbl_main_title = tk.Label(root, text="Motore ESG", font=(FONT_FAMILY, 18, "bold"), bg="#ffffff", fg="#2e7d32")
lbl_main_title.pack(pady=(18, 2))

lbl_main_sub = tk.Label(root, text="Estrazione Consumi Rapida", font=(FONT_FAMILY, 10), bg="#ffffff", fg="#64748b")
lbl_main_sub.pack(pady=(0, 16))

frame_inputs = tk.Frame(root, bg="#ffffff")
frame_inputs.pack(padx=30, fill="x", pady=5)

tk.Label(frame_inputs, text="Nome Azienda/Cliente (senza spazi):", font=(FONT_FAMILY, 10, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w", pady=(0, 4))
entry_azienda = tk.Entry(frame_inputs, font=(FONT_FAMILY, 10), relief="solid", bd=1)
entry_azienda.pack(fill="x", ipady=5, pady=(0, 12))

tk.Label(frame_inputs, text="Percorso Server/Cartella Principale (Root):", font=(FONT_FAMILY, 10, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w", pady=(0, 4))

frame_path_row = tk.Frame(frame_inputs, bg="#ffffff")
frame_path_row.pack(fill="x")

entry_path = tk.Entry(frame_path_row, font=(FONT_FAMILY, 10), relief="solid", bd=1)
entry_path.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 8))

btn_browse = tk.Button(frame_path_row, text="Sfoglia...", command=seleziona_cartella, font=(FONT_FAMILY, 9, "bold"), bg="#e2e8f0", fg="#334155", relief="flat", cursor="hand2", padx=16, pady=5)
btn_browse.pack(side="right")

frame_cards = tk.Frame(root, bg="#ffffff")
frame_cards.pack(padx=30, pady=20, fill="both", expand=True)

# Card 1: Energia Elettrica
card_elec = tk.Frame(frame_cards, bg="#ffffff", highlightbackground="#cbd5e1", highlightthickness=1, bd=0)
card_elec.pack(side="left", fill="both", expand=True, padx=(0, 10))

tk.Label(card_elec, text="Ambiente - Energia Elettrica", font=(FONT_FAMILY, 12, "bold"), bg="#ffffff", fg="#2e7d32").pack(pady=(16, 6))
tk.Label(card_elec, text="Estrai kWh da Prima Pagina\ncon Link diretti.", font=(FONT_FAMILY, 9), bg="#ffffff", fg="#64748b", justify="center").pack(pady=(0, 12))

btn_elec = tk.Button(card_elec, text="Avvia Energia Elettrica", command=avvia_energia_elettrica, bg="#2e7d32", fg="white", font=(FONT_FAMILY, 10, "bold"), relief="flat", cursor="hand2", pady=8, padx=12)
btn_elec.pack(pady=(0, 12))

progress_elec = ttk.Progressbar(card_elec, orient="horizontal", length=220, mode="determinate")
progress_elec.pack(pady=(0, 4))

lbl_status_elec = tk.Label(card_elec, text="In attesa...", font=(FONT_FAMILY, 8), bg="#ffffff", fg="#64748b")
lbl_status_elec.pack(pady=(0, 14))

# Card 2: Trasporti
card_trans = tk.Frame(frame_cards, bg="#ffffff", highlightbackground="#cbd5e1", highlightthickness=1, bd=0)
card_trans.pack(side="right", fill="both", expand=True, padx=(10, 0))

tk.Label(card_trans, text="Ambiente - Trasporti", font=(FONT_FAMILY, 12, "bold"), bg="#ffffff", fg="#2e7d32").pack(pady=(16, 6))
tk.Label(card_trans, text="Estrai litri e tipo carburante\n(gasolio/benzina).", font=(FONT_FAMILY, 9), bg="#ffffff", fg="#64748b", justify="center").pack(pady=(0, 12))

btn_trans = tk.Button(card_trans, text="Avvia Trasporti", command=avvia_trasporti, bg="#2e7d32", fg="white", font=(FONT_FAMILY, 10, "bold"), relief="flat", cursor="hand2", pady=8, padx=12)
btn_trans.pack(pady=(0, 12))

progress_trans = ttk.Progressbar(card_trans, orient="horizontal", length=220, mode="determinate")
progress_trans.pack(pady=(0, 4))

lbl_status_trans = tk.Label(card_trans, text="In attesa...", font=(FONT_FAMILY, 8), bg="#ffffff", fg="#64748b")
lbl_status_trans.pack(pady=(0, 14))

root.mainloop()
